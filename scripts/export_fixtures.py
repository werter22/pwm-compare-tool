import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

# -----------------------------
# Pfade (immer Repo-root stabil)
# -----------------------------
ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "scripts" / "input" / "Passwortmanagement_Kriterienkatalog.xlsx"

OUTDIR = ROOT / "public" / "fixtures"
OUT_TREE = OUTDIR / "tree.json"
OUT_PRODUCTS = OUTDIR / "products.json"
OUT_SCORES = OUTDIR / "scores.json"

# -----------------------------
# Domänen-Mapping (Excel -> UI)
# -----------------------------
DOMAIN_MAP = {
    "Security & Compliance": ("d1", "Sicherheit & Compliance"),
    "Datenhoheit, Lieferkette & Governance": ("d2", "Datenhoheit, Lieferkette & Governance"),
    "Produkt, Betrieb & Adoption": ("d3", "Produkt, Betrieb & Adoption"),
}

# Sheets, die KEINE Produkte sind
NON_PRODUCT_SHEETS = {
    "Kriterienkatalog",
    "Security & Compliance",
    "Datenhoheit, Lieferkette & Gove",
    "Produkt, Betrieb & Adoption_x0009_",
    "Vorlage",
    "Stammdaten",
    "Mastersheet",
}

CHAPTER_RE = re.compile(r"^\d+(\.\d+)+$")  # z.B. 4.1.2
CRIT_PREFIX_RE = re.compile(r"^(\d+(\.\d+)*)\s+")  # z.B. "3.2 Sicherheitsreife..." -> "3.2"


def is_empty(v: Any) -> bool:
    return v is None or str(v).strip() == "" or str(v).strip() == "\xa0"


def norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def slug(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    s = s.replace("&", " und ")
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def short_hash(s: str, n: int = 8) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()[:n]


def split_evidence(raw: Any) -> List[str]:
    if is_empty(raw):
        return []
    s = str(raw).strip()
    parts = re.split(r"[\n;]+", s)
    return [p.strip() for p in parts if p.strip()]


def parse_evidence_links(evid_raw: Any, evidenz_typ: Any = None) -> List[Dict[str, str]]:
    parts = split_evidence(evid_raw)
    if not parts:
        return []

    typ = "" if is_empty(evidenz_typ) else norm_text(str(evidenz_typ))

    links: List[Dict[str, str]] = []
    for i, p in enumerate(parts, start=1):
        m = re.search(r"(https?://\S+)", p)
        url = m.group(1).rstrip(").,") if m else ""

        label = f"Quelle {i}"
        if typ:
            label = f"{label} – {typ}"

        links.append({"label": label, "url": url})

    return links



# -----------------------------
# Robust Header Matching
# -----------------------------
def norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("–", "-").replace("—", "-")
    return s


def find_col(cols: Dict[str, int], *needles: str) -> Optional[int]:
    """
    Sucht eine Spalte anhand von Teilstrings im Header (case/space tolerant).
    """
    needles_n = [norm_header(n) for n in needles if n and n.strip()]
    for k, idx in cols.items():
        nk = norm_header(k)
        if any(n in nk for n in needles_n):
            return idx
    return None


def find_header_row(ws, max_rows=120, max_cols=80) -> Tuple[int, Dict[str, int]]:
    """
    Findet die Header-Zeile, indem sie nach einer Zeile sucht,
    die 'id'+'kapitel' und 'unterkriterium' (tolerant) enthält.
    """
    for r in range(1, max_rows + 1):
        vals = [ws.cell(r, c).value for c in range(1, max_cols + 1)]
        headers = [str(v).strip() for v in vals if isinstance(v, str) and v.strip()]
        hn = [norm_header(h) for h in headers]

        has_id_kapitel = any(("id" in h and "kapitel" in h) for h in hn)
        has_unterkriterium = any("unterkriterium" in h for h in hn)

        if has_id_kapitel and has_unterkriterium:
            col_map: Dict[str, int] = {}
            for c in range(1, max_cols + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip():
                    col_map[v.strip()] = c
            return r, col_map

    raise RuntimeError(f"Headerzeile nicht gefunden in Sheet '{ws.title}'")


def extract_criterion_prefix(criterion_name: str) -> str:
    s = norm_text(criterion_name)
    m = CRIT_PREFIX_RE.match(s)
    if m:
        return m.group(1)
    return slug(s)[:12] or "crit"


def clean_criterion_name(criterion_name: str) -> str:
    s = norm_text(criterion_name)
    return CRIT_PREFIX_RE.sub("", s).strip() or s


def make_subcriterion_id(domain_id: str, criterion_prefix: str, sub_name: str, chapter_ref: Optional[str]) -> str:
    """
    Eindeutige ID:
    - wenn Kapitelnummer vorhanden: s_<domain>_<kapitel>
    - sonst: s_<domain>_<critprefix>_<slug>_<hash>
    """
    if chapter_ref and CHAPTER_RE.match(chapter_ref):
        return f"s_{domain_id}_{chapter_ref.replace('.', '_')}"
    base = f"{domain_id}|{criterion_prefix}|{norm_text(sub_name)}"
    return f"s_{domain_id}_{criterion_prefix.replace('.', '_')}_{slug(sub_name)[:24]}_{short_hash(base)}"


@dataclass
class RowItem:
    domain_id: str
    domain_name: str
    criterion_id: str
    criterion_name: str
    criterion_prefix: str
    subcriterion_id: str
    subcriterion_name: str
    short_desc: str
    chapter_ref: Optional[str]


def iter_items(ws) -> List[RowItem]:
    """
    Parst ein Sheet (Vorlage oder Produkt) und extrahiert ALLE Unterkriterien,
    auch wenn 'ID (Kapitel)' bei Unterkriterien-Zeilen leer ist.
    """
    header_row, cols = find_header_row(ws)

    col_id = find_col(cols, "id", "kapitel")
    col_crit = find_col(cols, "kriterium")
    col_sub = find_col(cols, "unterkriterium")
    col_pruef = find_col(cols, "prüfverfahren", "pruefverfahren", "zusammenfassung")

    if not col_sub:
        raise RuntimeError(f"Unterkriterium-Spalte fehlt in '{ws.title}'")

    current_domain_id: Optional[str] = None
    current_domain_name: Optional[str] = None
    current_criterion_name: Optional[str] = None
    current_criterion_prefix: Optional[str] = None
    current_criterion_id: Optional[str] = None

    items: List[RowItem] = []

    for r in range(1, ws.max_row + 1):
        # Domain-Erkennung über Spalte A
        a = ws.cell(r, 1).value
        if isinstance(a, str) and a.strip() in DOMAIN_MAP:
            did, dname = DOMAIN_MAP[a.strip()]
            current_domain_id, current_domain_name = did, dname
            current_criterion_name = None
            current_criterion_prefix = None
            current_criterion_id = None

        if r <= header_row:
            continue

        sub = ws.cell(r, col_sub).value if col_sub else None
        if is_empty(sub):
            continue

        if not current_domain_id or not current_domain_name:
            continue

        # Criterion-Kontext: wenn Kriterium befüllt, aktualisieren; sonst beibehalten
        crit = ws.cell(r, col_crit).value if col_crit else None
        if not is_empty(crit):
            current_criterion_name = norm_text(str(crit))
            current_criterion_prefix = extract_criterion_prefix(current_criterion_name)
            current_criterion_id = f"c_{current_domain_id}_{current_criterion_prefix.replace('.', '_')}"

        if not current_criterion_name or not current_criterion_prefix or not current_criterion_id:
            continue

        sub_name = norm_text(str(sub))

        chap_raw = ws.cell(r, col_id).value if col_id else None
        chap = norm_text(str(chap_raw)) if not is_empty(chap_raw) else None
        chapter_ref = chap if chap else None

        short_desc = ""
        if col_pruef:
            pv = ws.cell(r, col_pruef).value
            if not is_empty(pv):
                txt = norm_text(str(pv))
                short_desc = txt[:180] + ("…" if len(txt) > 180 else "")

        sub_id = make_subcriterion_id(
            current_domain_id,
            current_criterion_prefix,
            sub_name,
            chap if chap and CHAPTER_RE.match(chap) else None,
        )

        items.append(
            RowItem(
                domain_id=current_domain_id,
                domain_name=current_domain_name,
                criterion_id=current_criterion_id,
                criterion_name=clean_criterion_name(current_criterion_name),
                criterion_prefix=current_criterion_prefix,
                subcriterion_id=sub_id,
                subcriterion_name=sub_name,
                short_desc=short_desc,
                chapter_ref=chapter_ref,
            )
        )

    return items


def build_tree_from_template(ws_template) -> Dict[str, Any]:
    items = iter_items(ws_template)

    domains: Dict[str, Dict[str, Any]] = {}
    criteria_index: Dict[Tuple[str, str], Dict[str, Any]] = {}

    for it in items:
        dnode = domains.get(it.domain_id)
        if not dnode:
            dnode = {"id": it.domain_id, "name": it.domain_name, "criteria": []}
            domains[it.domain_id] = dnode

        ckey = (it.domain_id, it.criterion_id)
        cnode = criteria_index.get(ckey)
        if not cnode:
            cnode = {"id": it.criterion_id, "name": it.criterion_name, "subcriteria": []}
            criteria_index[ckey] = cnode
            dnode["criteria"].append(cnode)

        if not any(s["id"] == it.subcriterion_id for s in cnode["subcriteria"]):
            sc = {"id": it.subcriterion_id, "name": it.subcriterion_name, "short_desc": it.short_desc}
            if it.chapter_ref:
                sc["chapter_ref"] = it.chapter_ref
            cnode["subcriteria"].append(sc)

    ordered = []
    for did in ["d1", "d2", "d3"]:
        if did in domains:
            ordered.append(domains[did])

    return {"domains": ordered}


def parse_product_scores(ws_product, template_items: List[RowItem], product_id: str) -> List[Dict[str, Any]]:
    """
    Liest Scores aus Produkt-Sheet. Ordnet Unterkriterien über die Reihenfolge
    der Unterkriterium-Zeilen zur Vorlage zu (robust bei leeren Kapitel-IDs).
    """
    header_row, cols = find_header_row(ws_product)

    col_score = find_col(cols, "scoring", "score", "bewertung")
    col_comment = find_col(cols, "kommentar", "kurzbefund", "comment")
    col_evid = find_col(cols, "evidenz", "quelle", "source", "link")
    col_evid_typ = find_col(cols, "evidenz-typ", "evidenz typ", "evidenztyp", "evidence typ")

    col_sub = find_col(cols, "unterkriterium")

    if not col_score:
        print("Gefundene Header in", ws_product.title, ":", list(cols.keys()))
        raise RuntimeError(f"Scoring-Spalte fehlt in '{ws_product.title}'")
    if not col_sub:
        print("Gefundene Header in", ws_product.title, ":", list(cols.keys()))
        raise RuntimeError(f"Unterkriterium-Spalte fehlt in '{ws_product.title}'")

    # Vorlage + Produkt zu Items parsen
    prod_items = iter_items(ws_product)

    if len(prod_items) != len(template_items):
        print(f"WARN: '{ws_product.title}' hat {len(prod_items)} Unterkriterien, Vorlage hat {len(template_items)}.")

    n = min(len(prod_items), len(template_items))
    out: List[Dict[str, Any]] = []

    # Zeilennummern der Unterkriterium-Zeilen sammeln
    current_domain_seen = False
    rows_for_subcriteria: List[int] = []
    for r in range(1, ws_product.max_row + 1):
        a = ws_product.cell(r, 1).value
        if isinstance(a, str) and a.strip() in DOMAIN_MAP:
            current_domain_seen = True
        if r <= header_row:
            continue
        if not current_domain_seen:
            continue
        sub = ws_product.cell(r, col_sub).value
        if is_empty(sub):
            continue
        rows_for_subcriteria.append(r)

    if len(rows_for_subcriteria) < n:
        print(f"WARN: Zeilenerkennung '{ws_product.title}' liefert weniger Rows ({len(rows_for_subcriteria)}) als Items ({n}).")
        n = min(n, len(rows_for_subcriteria))

    for i in range(n):
        tpl = template_items[i]
        r = rows_for_subcriteria[i]

        raw = ws_product.cell(r, col_score).value
        score = 0
        if not is_empty(raw):
            try:
                score = int(raw)
            except Exception:
                score = 0
        if score not in (0, 1, 2):
            score = 0

        comment = ""
        if col_comment:
            v = ws_product.cell(r, col_comment).value
            if not is_empty(v):
                comment = norm_text(str(v))

        evid = ws_product.cell(r, col_evid).value if col_evid else None
        evid_typ = ws_product.cell(r, col_evid_typ).value if col_evid_typ else None
        evidenz_links = parse_evidence_links(evid, evid_typ)

        out.append(
            {
                "product_id": product_id,
                "subcriterion_id": tpl.subcriterion_id,
                "score": score,
                "audit_comment": comment,
                "evidenz_links": evidenz_links,
            }
        )

    return out


def main():
    if not INPUT.exists():
        raise FileNotFoundError(f"Excel nicht gefunden: {INPUT}")

    wb = load_workbook(INPUT, data_only=True)

    if "Vorlage" not in wb.sheetnames:
        raise RuntimeError("Sheet 'Vorlage' nicht gefunden. Bitte sicherstellen, dass die Excel eine 'Vorlage' enthält.")
    ws_template = wb["Vorlage"]

    template_items = iter_items(ws_template)
    tree = build_tree_from_template(ws_template)

    # Produkte bestimmen (alle Sheets ausser NON_PRODUCT)
    product_sheetnames = [s for s in wb.sheetnames if s not in NON_PRODUCT_SHEETS and s != "Vorlage"]
    products = [{"id": slug(s), "name": s, "description": ""} for s in product_sheetnames]

    scores: List[Dict[str, Any]] = []
    for s in product_sheetnames:
        pid = slug(s)
        ws = wb[s]
        scores.extend(parse_product_scores(ws, template_items, pid))

    # Duplikate im Tree prüfen
    all_ids = []
    for d in tree["domains"]:
        for c in d["criteria"]:
            for sc in c["subcriteria"]:
                all_ids.append(sc["id"])
    dup = {x for x in all_ids if all_ids.count(x) > 1}
    if dup:
        print("WARN: Duplizierte subcriterion_id im Tree (Sample):", sorted(list(dup))[:20])

    OUTDIR.mkdir(parents=True, exist_ok=True)
    OUT_TREE.write_text(json.dumps(tree, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_PRODUCTS.write_text(json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_SCORES.write_text(json.dumps(scores, ensure_ascii=False, indent=2), encoding="utf-8")

    print("OK: exported fixtures")
    print(f"- {OUT_TREE}")
    print(f"- {OUT_PRODUCTS}")
    print(f"- {OUT_SCORES}")
    print(f"Produkte: {len(products)} | Unterkriterien (Vorlage): {len(template_items)} | Score-Zeilen: {len(scores)}")


if __name__ == "__main__":
    main()
